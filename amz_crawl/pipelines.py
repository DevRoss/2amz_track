# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html

from scrapy.utils.project import get_project_settings
from scrapy.pipelines.images import ImagesPipeline
from items import AmazonDepartmentItem
import os
import platform
from datetime import datetime
from openpyxl import Workbook

settings = get_project_settings()
all_resource_path = settings.attributes.get('RESOURCE_STORE').value
resource_path = settings.attributes.get('IMAGES_STORE').value
today_path = os.path.join(all_resource_path, datetime.now().strftime('%y-%m-%d'))

import csv

from scrapy import signals
from scrapy.contrib.exporter import CsvItemExporter


class CSVPipeline(object):
    def __init__(self):
        self.files = {}

    @classmethod
    def from_crawler(cls, crawler):
        pipeline = cls()
        crawler.signals.connect(pipeline.spider_opened, signals.spider_opened)
        crawler.signals.connect(pipeline.spider_closed, signals.spider_closed)
        return pipeline

    def spider_opened(self, spider):
        pass

    def spider_closed(self, spider):
        if self.exporter:
            self.exporter.finish_exporting()
            file = self.files.pop(spider)
            file.close()

    def process_item(self, item, spider):
        file_path = os.path.join(resource_path, 'category.csv')
        if not os.path.exists(file_path):
            file = open(file_path, 'w+b')
            self.files[spider] = file
            self.exporter = CsvItemExporter(file)
            self.exporter.fields_to_export = ['belong', 'title', 'url', 'category', 'sub_category']
            self.exporter.start_exporting()
            self.exporter.export_item(item)
        else:
            self.exporter.export_item(item)

        return item


class AmzCrawlXLSXPipeline(object):
    wb = Workbook()

    def process_item(self, item, spider):  # 工序具体内容
        self.insert_item_to_xlsx(item)
        return item

    def insert_item_to_xlsx(self, item):
        category = item['category']
        line = [category, item['sub_category'], item['title'], item['rank']]  # 把数据中每一项整理出来
        belong_dir = os.path.join(today_path, item['belong'])
        if not os.path.exists(belong_dir):
            os.makedirs(belong_dir)

        xlsx_path = os.path.join(today_path, datetime.now().strftime('%y-%m-%d') + '.xlsx')

        ws = self.wb.active

        if category in self.wb.sheetnames:
            ws = self.wb.get_sheet_by_name(category)
            ws.append(line)

        if category not in self.wb.sheetnames:
            if ws.title == 'Sheet':
                ws.title = category
            else:
                ws = self.wb.create_sheet(title=category)
            ws.append(['分类', '子分类', '标题', '排名'])  # 设置表头
            ws.append(line)  # 将数据以行的形式添加到xlsx中

            self.wb.save(xlsx_path)  # 保存xlsx文件


class AmzCrawlPipeline(object):
    def process_item(self, item, spider):
        self.process_file(item, spider)
        return item

    # 文件夹
    def process_file(self, item, spider):
        today_dir = os.path.join(resource_path, datetime.now().strftime('%y-%m-%d'))
        belong_dir = os.path.join(today_dir, item['belong'])
        sub_category_dir = os.path.join(belong_dir, item['sub_category'])

        if not os.path.exists(sub_category_dir):
            os.makedirs(sub_category_dir, exist_ok=True)


class AmazonOrdersImagePipeline(ImagesPipeline):
    def item_completed(self, results, item, info):
        if "front_image_url" in item:
            for ok, value in results:
                if platform.system() == 'Windows':  # full 图片位置
                    image_file_path = value["path"].replace('/', '\\')
                else:
                    image_file_path = value["path"]

                old_path = os.path.join(resource_path, image_file_path)

                if os.path.exists(old_path):
                    self.move_old_new(image_file_path, item)

        return item

    def move_old_new(self, image_file_path, item):

        today_dir = os.path.join(resource_path, datetime.now().strftime('%y-%m-%d'))
        belong_dir = os.path.join(today_dir, item['belong'])
        sub_category_dir = os.path.join(belong_dir, item['sub_category'])

        if not os.path.exists(sub_category_dir):
            os.makedirs(sub_category_dir, exist_ok=True)
        new_path = os.path.join(sub_category_dir, item['rank'] + '.' + item['des'] + '.jpg')
        old_path = os.path.join(resource_path, image_file_path)

        return new_path, old_path

        # from sqlalchemy.orm import sessionmaker
        # from amz_crawl.items import db_connect
        #
        # class AmzDataBasePipeline(object):
        #     """保存文章到数据库"""
        #     def __init__(self):
        #         engine = db_connect()
        #         create_news_table(engine)
        #         self.Session = sessionmaker(bind=engine)
        #     def open_spider(self, spider):
        #         """This method is called when the spider is opened."""
        #         pass
        #     def process_item(self, item, spider):
        #         a = Article(url=item["url"],
        #                     title=item["title"].encode("utf-8"),
        #                     publish_time=item["publish_time"].encode("utf-8"),
        #                     body=item["body"].encode("utf-8"),
        #                     source_site=item["source_site"].encode("utf-8"))
        #         with session_scope(self.Session) as session:
        #             session.add(a)
        #     def close_spider(self, spider):
        #         pass
